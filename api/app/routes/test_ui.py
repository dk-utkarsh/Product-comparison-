"""
Test UI routes. NOT part of the production API surface — these exist so
you can drop an Excel of Dentalkart products into a browser and see the
matching engine work end-to-end without writing client code.

Routes:
  POST /test/upload-catalog   ingest xlsx into dentalkart_catalog
  POST /test/search           top-K nearest neighbours for a query
  POST /test/run-batch        xlsx of queries -> top-K per row
  GET  /test/catalog-count    rows currently in the catalog
  POST /test/truncate-catalog wipe the catalog (for re-uploading)
"""
from __future__ import annotations

import io
from typing import Any

import openpyxl
from fastapi import APIRouter, File, HTTPException, UploadFile
from pydantic import BaseModel, Field

from app.db import get_db
from app.matching.index import CatalogIndex
from app.matching.triage import triage
from app.schemas import RankedCandidate
from scripts.build_catalog_index import ingest_rows

router = APIRouter(prefix="/test", tags=["test-ui"])


_NAME_HEADERS = {"product name", "name", "product", "title"}
_SKU_HEADERS = {"sku", "code", "item code", "product code"}
_BRAND_HEADERS = {"brand", "manufacturer"}


def _parse_xlsx_to_rows(content: bytes) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    iter_rows = ws.iter_rows(values_only=True)
    try:
        header = next(iter_rows)
    except StopIteration:
        return []

    headers_lc = [str(h).strip().lower() if h is not None else "" for h in header]

    def find_idx(candidates: set[str]) -> int | None:
        for i, h in enumerate(headers_lc):
            if h in candidates:
                return i
        return None

    name_idx = find_idx(_NAME_HEADERS)
    if name_idx is None:
        # If there's a single column, assume it's the name.
        if len(headers_lc) == 1:
            name_idx = 0
        else:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"could not find a name column. Looked for one of {sorted(_NAME_HEADERS)}; "
                    f"got headers {headers_lc}"
                ),
            )
    sku_idx = find_idx(_SKU_HEADERS)
    brand_idx = find_idx(_BRAND_HEADERS)

    rows: list[dict[str, str]] = []
    for r in iter_rows:
        if not r or r[name_idx] is None:
            continue
        row = {"name": str(r[name_idx]).strip()}
        if sku_idx is not None and r[sku_idx] is not None:
            row["sku"] = str(r[sku_idx]).strip()
        if brand_idx is not None and r[brand_idx] is not None:
            row["brand"] = str(r[brand_idx]).strip().lower()
        if row["name"]:
            rows.append(row)
    return rows


class UploadResult(BaseModel):
    inserted: int
    total_in_catalog: int
    detected_columns: dict[str, str | None]


@router.post("/upload-catalog", response_model=UploadResult)
async def upload_catalog(file: UploadFile = File(...)) -> UploadResult:
    content = await file.read()
    rows = _parse_xlsx_to_rows(content)
    if not rows:
        raise HTTPException(status_code=400, detail="no usable rows found in the xlsx")

    db = await get_db()
    try:
        inserted = await ingest_rows(rows, db)
        total = await db.fetchrow("SELECT count(*) AS c FROM dentalkart_catalog")
        first = rows[0]
        detected = {
            "name": "name" if "name" in first else None,
            "sku": "sku" if "sku" in first else None,
            "brand": "brand" if "brand" in first else None,
        }
        return UploadResult(
            inserted=inserted,
            total_in_catalog=int(total["c"]) if total else 0,
            detected_columns=detected,
        )
    finally:
        await db.close()


class SearchRequest(BaseModel):
    search: str = Field(min_length=1)
    k: int = Field(default=5, ge=1, le=50)
    run_triage: bool = Field(default=True)


class SearchHit(BaseModel):
    name: str
    sku: str | None
    brand: str | None
    cosine: float
    triage: RankedCandidate | None = None


class SearchResponse(BaseModel):
    query: str
    hits: list[SearchHit]


@router.post("/search", response_model=SearchResponse)
async def search(req: SearchRequest) -> SearchResponse:
    db = await get_db()
    try:
        idx = CatalogIndex(db)
        hits = await idx.top_k(req.search, k=req.k)
        out: list[SearchHit] = []
        for h in hits:
            tri: RankedCandidate | None = None
            if req.run_triage:
                r = triage(req.search, h.name)
                tri = RankedCandidate(
                    candidate=h.name,
                    verdict=r.verdict.value,
                    score=r.score,
                    cosine=r.cosine,
                    reasons=r.reasons,
                )
            out.append(
                SearchHit(
                    name=h.name,
                    sku=h.sku,
                    brand=h.brand,
                    cosine=h.cosine,
                    triage=tri,
                )
            )
        return SearchResponse(query=req.search, hits=out)
    finally:
        await db.close()


class BatchRunResult(BaseModel):
    query: str
    hits: list[SearchHit]


class BatchResponse(BaseModel):
    total_queries: int
    results: list[BatchRunResult]


@router.post("/run-batch", response_model=BatchResponse)
async def run_batch(
    file: UploadFile = File(...),
    k: int = 5,
) -> BatchResponse:
    content = await file.read()
    rows = _parse_xlsx_to_rows(content)
    if not rows:
        raise HTTPException(status_code=400, detail="no usable rows found")

    db = await get_db()
    try:
        idx = CatalogIndex(db)
        results: list[BatchRunResult] = []
        for r in rows:
            q = r["name"]
            hits = await idx.top_k(q, k=k)
            sh: list[SearchHit] = []
            for h in hits:
                t = triage(q, h.name)
                sh.append(
                    SearchHit(
                        name=h.name,
                        sku=h.sku,
                        brand=h.brand,
                        cosine=h.cosine,
                        triage=RankedCandidate(
                            candidate=h.name,
                            verdict=t.verdict.value,
                            score=t.score,
                            cosine=t.cosine,
                            reasons=t.reasons,
                        ),
                    )
                )
            results.append(BatchRunResult(query=q, hits=sh))
        return BatchResponse(total_queries=len(results), results=results)
    finally:
        await db.close()


class CatalogCount(BaseModel):
    count: int


@router.get("/catalog-count", response_model=CatalogCount)
async def catalog_count() -> CatalogCount:
    db = await get_db()
    try:
        row = await db.fetchrow("SELECT count(*) AS c FROM dentalkart_catalog")
        return CatalogCount(count=int(row["c"]) if row else 0)
    finally:
        await db.close()


@router.post("/truncate-catalog")
async def truncate_catalog() -> dict[str, Any]:
    db = await get_db()
    try:
        await db.execute("DELETE FROM dentalkart_catalog")
        return {"status": "ok", "count": 0}
    finally:
        await db.close()
