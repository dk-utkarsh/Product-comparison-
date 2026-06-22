"""
End-to-end comparison routes — Python orchestrates, TS scrapers fetch.

For each Dentalkart product (uploaded via xlsx):
  1. Fan out to every competitor via the TS scraper bridge in parallel.
  2. Each competitor returns up to N candidate listings.
  3. Run every (dk_name, candidate_name) pair through Python triage.
  4. Keep the highest-scoring confirmed/possible candidate per competitor.
  5. Compute Δ vs Dentalkart price.

UI hits POST /compare for one row or POST /compare-batch for an xlsx.
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
import re
from collections.abc import AsyncIterator

import openpyxl
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from app import pipeline, registry
from app.db import Database, get_db
from app.matching.llm_judge import JudgeBudget
from app.matching.query_builder import ProductContext, extract_smart_queries
from app.matching.score import Verdict
from app.matching.attributes import extract_attributes
from app.matching.gates import gate_check
from app.matching.normalize import normalize_for_match
from app.matching.structured import ProductRecord
from app.matching.tokens import distinguishing_tokens, fuzz_ratio
from app.matching.triage import triage_batch
from app.matching.variant_spec import (
    VariantSpec,
    base_name,
    base_quantity,
    config_from_text,
)
from app.scrapers.bridge import (
    COMPETITORS,
    CompetitorProduct,
    fetch_product,
    scrape_competitor,
)
from app.settings import get_settings

router = APIRouter(prefix="/compare", tags=["compare"])

_ACCEPT_VERDICTS = (Verdict.CONFIRMED.value, Verdict.POSSIBLE.value)


class DkRow(BaseModel):
    name: str = Field(min_length=1)


class CompetitorMatch(BaseModel):
    competitor_id: str
    competitor_name: str
    candidates_seen: int
    matched_name: str | None
    matched_url: str | None
    matched_price: float | None
    matched_image: str | None
    in_stock: bool | None
    verdict: str | None
    score: float | None
    cosine: float | None
    reasons: list[str] = []
    price_diff_vs_dk: float | None = None
    price_diff_per_unit: float | None = None
    matched_by: str | None = None
    pack_note: str | None = None
    spec_match: str | None = None  # exact | same-tier | different-size


class CompareResult(BaseModel):
    dentalkart: DkRow
    dentalkart_match: CompetitorMatch | None = None
    competitors: list[CompetitorMatch]


class CompareBatchResponse(BaseModel):
    total: int
    results: list[CompareResult]


def _prefilter_candidates(
    search: str, candidates: list[CompetitorProduct]
) -> list[CompetitorProduct]:
    """Sparse pre-filter: only keep candidates that share at least one
    distinguishing token with the search (stopwords excluded).

    Without this, broader fallback queries dump a lot of obviously-unrelated
    products into the pool and waste embedder cycles. With it, the pool
    shrinks to plausible candidates only.
    """
    sig = distinguishing_tokens(search)
    if not sig:
        return candidates
    kept: list[CompetitorProduct] = []
    for c in candidates:
        if not c.name:
            continue
        if distinguishing_tokens(c.name) & sig:
            kept.append(c)
    return kept


def _in_price_band(
    cand_price: float, dk_price: float | None, max_ratio: float
) -> bool:
    """True when the competitor price is plausibly the same product as DK.

    If we don't know the DK price (e.g. dentalkart.com had no match for
    the row), we can't check price sanity — pass everything through.
    """
    if not dk_price or dk_price <= 0 or cand_price <= 0:
        return True
    ratio = cand_price / dk_price
    return (1.0 / max_ratio) <= ratio <= max_ratio


# Specialization markers — a candidate carrying one of these that the user's
# input does NOT is a more specific sub-variant (Capsules / Extra / Pack-of-N /
# refill / combo / Set-of-N / mini / drills). When the input is ambiguous we
# should prefer the plain base product over such a specialization.
_QUALIFIER_PATTERNS: tuple[re.Pattern[str], ...] = (
    re.compile(r"\bextra\b", re.I),
    re.compile(r"\bcapsules?\b", re.I),
    re.compile(r"\brefills?\b", re.I),
    re.compile(r"\bcombo\b", re.I),
    re.compile(r"\bpack\s+of\s+\d+\b", re.I),
    re.compile(r"\bset\s+of\s+\d+\b", re.I),
    re.compile(r"\b(mini|trial|sample)\b", re.I),
    re.compile(r"\bonly\b", re.I),
    re.compile(r"\bdrills?\b", re.I),
)


def _code_match_bonus(input_name: str, cand_name: str) -> float:
    """Ranking boost for a DK candidate that shares a distinctive model/size code
    with the input (suture '#2-0', model 'DL-300', dims '17x25') — so the
    exact-size product wins over a size-less sibling (e.g. '…Silk Suture Reels')."""
    in_codes = set(extract_attributes(normalize_for_match(input_name)).model_codes)
    if not in_codes:
        return 0.0
    cand_codes = set(extract_attributes(normalize_for_match(cand_name)).model_codes)
    return 0.25 if in_codes & cand_codes else 0.0


def _qualifier_penalty(input_name: str, cand_name: str) -> float:
    """Ranking demotion for a candidate that introduces specialization markers
    absent from the user's input — so "GC Gold Label 9" prefers the base
    Posterior Restorative over "...Extra Capsules Pack Of 30". Only relative
    ranking is affected; the displayed score/verdict is unchanged."""
    pen = 0.0
    for rx in _QUALIFIER_PATTERNS:
        if rx.search(cand_name) and not rx.search(input_name):
            pen += 0.12
    return min(pen, 0.30)


def _best_match(
    dk_name: str,
    candidates: list[CompetitorProduct],
    dk_price: float | None,
    qualifier_ref: str | None = None,
) -> CompetitorMatch | None:
    """Pre-filter, batch-triage, pick the highest-scoring confirmed/possible
    candidate that also lies within the DK price band. `qualifier_ref` (the
    user's original input) demotes more-specific sub-variants when ambiguous."""
    if not candidates:
        return None

    pool = [c for c in candidates if c.name and c.price > 0]
    pool = _prefilter_candidates(dk_name, pool)
    if not pool:
        return None

    results = triage_batch(dk_name, [c.name for c in pool])
    max_ratio = get_settings().price_band_max_ratio
    ref = qualifier_ref or dk_name

    best_score = -1.0
    best_cand: CompetitorProduct | None = None
    best_result = None
    for cand, r in zip(pool, results, strict=True):
        if r.verdict.value not in _ACCEPT_VERDICTS:
            continue
        if not _in_price_band(cand.price, dk_price, max_ratio):
            # Cosine + token may agree, but the price says it's a
            # different product (a part vs the machine, a kit vs a single
            # instrument, etc.). Drop it.
            continue
        eff = (
            r.score
            - _qualifier_penalty(ref, cand.name)
            + _code_match_bonus(ref, cand.name)
        )
        if eff > best_score:
            best_score = eff
            best_cand = cand
            best_result = r

    if best_cand is None or best_result is None:
        return None

    diff: float | None = None
    if dk_price and dk_price > 0:
        diff = round(dk_price - best_cand.price, 2)

    return CompetitorMatch(
        competitor_id="",
        competitor_name="",
        candidates_seen=len(candidates),
        matched_name=best_cand.name,
        matched_url=best_cand.url,
        matched_price=best_cand.price,
        matched_image=best_cand.image,
        in_stock=best_cand.in_stock,
        verdict=best_result.verdict.value,
        score=best_result.score,
        cosine=best_result.cosine,
        reasons=best_result.reasons,
        price_diff_vs_dk=diff,
    )


def _empty_cell(cid: str, cname: str, seen: int) -> CompetitorMatch:
    return CompetitorMatch(
        competitor_id=cid,
        competitor_name=cname,
        candidates_seen=seen,
        matched_name=None,
        matched_url=None,
        matched_price=None,
        matched_image=None,
        in_stock=None,
        verdict=None,
        score=None,
        cosine=None,
        reasons=[],
        price_diff_vs_dk=None,
    )


def _per_unit_diff(
    dk_record: ProductRecord | None, c: CompetitorProduct,
    dk_price: float | None, dk_unit_price: float | None,
) -> float | None:
    """Per-unit Δ for an apples-to-apples comparison when the two sides differ in
    size. Prefer composition base quantity (e.g. ₹ per gram of powder); fall
    back to pack-size unit price."""
    c_spec = VariantSpec.from_dict(c.variant_spec)
    dk_spec = dk_record.variant_spec if dk_record else None
    if (
        dk_spec is not None and c_spec is not None
        and dk_price and dk_price > 0 and c.price > 0
    ):
        dk_qty, dk_unit = base_quantity(dk_spec)
        c_qty, c_unit = base_quantity(c_spec)
        if dk_unit == c_unit and dk_qty > 0 and c_qty > 0 and (dk_qty != c_qty):
            return round(dk_price / dk_qty - c.price / c_qty, 2)
    if dk_unit_price and dk_unit_price > 0 and c.unit_price > 0:
        return round(dk_unit_price - c.unit_price, 2)
    return None


def _cell_to_match(cid: str, cname: str, cell: pipeline.Cell,
                   dk_price: float | None,
                   dk_unit_price: float | None,
                   dk_record: ProductRecord | None = None) -> CompetitorMatch:
    c = cell.candidate
    if c is None or cell.verdict is None:
        return _empty_cell(cid, cname, cell.candidates_seen)
    diff = round(dk_price - c.price, 2) if dk_price and dk_price > 0 else None
    # Different size/pack makes the headline Δ misleading — also expose a
    # per-unit Δ so the UI can show an apples-to-apples comparison.
    unit_diff: float | None = None
    if cell.pack_note:
        unit_diff = _per_unit_diff(dk_record, c, dk_price, dk_unit_price)
    reasons = list(cell.reasons)
    # Show the correct product even when it's out of stock (flag it).
    if c.in_stock is False:
        reasons = ["out of stock", *reasons]
    return CompetitorMatch(
        competitor_id=cid, competitor_name=cname,
        candidates_seen=cell.candidates_seen,
        matched_name=c.name, matched_url=c.url, matched_price=c.price,
        matched_image=c.image, in_stock=c.in_stock,
        verdict=cell.verdict, score=cell.confidence, cosine=None,
        reasons=reasons, price_diff_vs_dk=diff,
        price_diff_per_unit=unit_diff,
        matched_by=cell.matched_by, pack_note=cell.pack_note,
        spec_match=cell.spec_match,
    )


def _paren_code(name: str) -> str:
    """Last parenthetical code of a product name, normalized — the per-variant
    model code that distinguishes siblings, e.g. "(KGF 8)" → "kgf 8",
    "(KO 12K P03A)" → "ko 12k p03a". (For grouped products whose children all
    share one code like "(JULL-DENT 223)" it simply doesn't narrow anything.)"""
    codes = re.findall(r"\(([^)]+)\)", name)
    if not codes:
        return ""
    return re.sub(r"\s+", " ", codes[-1]).strip().lower()


def _word_tokens(s: str) -> set[str]:
    return set(re.findall(r"[a-z0-9]+", s.lower()))


def _pick_dk_child(
    input_name: str, parent_name: str, variants: list[dict]
) -> dict | None:
    """For a grouped Dentalkart product, resolve the input/xlsx name to the
    specific child sub-variant: by config (tier/torque/Extra), then exact model
    code, then — when neither applies — by name when the input clearly names a
    child (e.g. 'Size 19'). Returns None when the input doesn't pin a child
    (keep the grouped listing as-is)."""
    if not variants:
        return None
    in_kit, in_torque, in_extra = config_from_text(input_name)
    in_code = _paren_code(input_name)
    # The shared parent code (e.g. "JULL-DENT 223") isn't a per-variant
    # discriminator; treat a code as useful only if it differs across children.
    distinct_codes = {_paren_code(str(v.get("name", ""))) for v in variants}
    # Only a real SKU/serial parenthetical counts as a child discriminator — a
    # descriptor like "(Pack of 5)" must NOT drill into an arbitrary length child
    # (Surgident GBR Screw ∅ 1.4mm → keep the base, don't pick "x 3mm").
    use_code = _looks_like_code(in_code) and len(distinct_codes - {""}) > 1
    in_norm = normalize_for_match(input_name)

    def best_by_name(pool: list[dict]) -> dict:
        return max(
            pool,
            key=lambda v: fuzz_ratio(in_norm, normalize_for_match(str(v.get("name", "")))),
        )

    if in_kit or in_torque or in_extra or use_code:
        compatible: list[dict] = []
        for v in variants:
            vs = VariantSpec.from_dict(v.get("variantSpec"))
            vk = vs.kit_tier if vs else None
            vt = vs.torque if vs else None
            ve = vs.is_extra if vs else False
            if in_kit and vk and vk != in_kit:
                continue
            if in_torque and vt and vt != in_torque:
                continue
            if in_extra != ve and (in_extra or ve):
                continue
            compatible.append(v)
        if not compatible:
            return None
        # Exact model-code match wins outright (KGF 8 vs KGF 9 vs KO 1/2).
        if use_code:
            coded = [v for v in compatible if _paren_code(str(v.get("name", ""))) == in_code]
            if coded:
                compatible = coded
        return best_by_name(compatible)

    # No config/code signal: resolve by name only when the input UNIQUELY names a
    # child. Compare against the tokens the input adds beyond the parent (e.g.
    # "Size 19" → {size,19,…}); the child sharing the most of them must be a
    # strict winner. Otherwise the input is ambiguous → keep the parent.
    extra = _word_tokens(input_name) - _word_tokens(parent_name)
    if not extra:
        return None
    counts = [(v, len(_word_tokens(str(v.get("name", ""))) & extra)) for v in variants]
    top = max(c for _, c in counts)
    if top == 0:
        return None
    tied = [v for v, c in counts if c == top]
    if len(tied) == 1:
        return tied[0]
    # Several children share the same input tokens — e.g. "#80" matches both the
    # "#80" child and the "#45-80" range child. Break the tie by name fuzz so the
    # exact size wins; keep the parent only if there's no strict winner.
    tied.sort(
        key=lambda v: fuzz_ratio(in_norm, normalize_for_match(str(v.get("name", "")))),
        reverse=True,
    )
    f0 = fuzz_ratio(in_norm, normalize_for_match(str(tied[0].get("name", ""))))
    f1 = fuzz_ratio(in_norm, normalize_for_match(str(tied[1].get("name", ""))))
    return tied[0] if f0 > f1 else None


_CODE_HAS_DIGIT = re.compile(r"\d")
_CODE_NONCODE = re.compile(
    r"\b(pack|set|sheets?|of|coarse|medium|fine|regular|assorted|"
    r"cm|mm|ml|mg|kg|gm?|microns?|µ|inch|inches|oz)\b",
    re.I,
)


def _looks_like_code(s: str) -> bool:
    """True for a SKU/serial-ish parenthetical (e.g. '5527/002/E', 'KGF 9',
    'S5083') — has a digit, short, and not a descriptor/measurement
    ('11.5cm', 'Pack of 300 sheets', 'Coarse/Medium')."""
    return (
        bool(s)
        and len(s) <= 24
        and _CODE_HAS_DIGIT.search(s) is not None
        and _CODE_NONCODE.search(s) is None
    )


def _build_dk_result(
    rich: CompetitorProduct, child: dict | None, input_name: str
) -> tuple[CompetitorMatch, ProductRecord]:
    """Build the Dentalkart self-match (CompetitorMatch + anchor ProductRecord)
    from a resolved product (`rich`) or one of its grouped children. Scores the
    match against the input name so an exact resolution reads confirmed/high."""
    if child is not None:
        name = str(child.get("name") or rich.name)
        price = float(child.get("price") or 0)
        in_stock = bool(child.get("inStock", rich.in_stock))
        pack_size = int(child.get("packSize") or 1)
        unit_price = float(child.get("unitPrice") or price)
        vspec = VariantSpec.from_dict(child.get("variantSpec"))
        image = str(child.get("image") or rich.image)
    else:
        name, price, in_stock = rich.name, rich.price, rich.in_stock
        pack_size = rich.pack_size
        unit_price = rich.unit_price or rich.price
        vspec = VariantSpec.from_dict(rich.variant_spec)
        image = rich.image

    tri = triage_batch(input_name, [name])
    dk_match = CompetitorMatch(
        competitor_id="dentalkart", competitor_name="Dentalkart",
        candidates_seen=0, matched_name=name, matched_url=rich.url,
        matched_price=price, matched_image=image, in_stock=in_stock,
        verdict=tri[0].verdict.value if tri else None,
        score=tri[0].score if tri else None,
        cosine=tri[0].cosine if tri else None, reasons=[],
    )
    dk_record = ProductRecord(
        name=name, url=rich.url, description=rich.description,
        packaging=rich.packaging, price=price, mrp=rich.mrp,
        pack_size=pack_size, unit_price=unit_price, sku=rich.sku,
        source="dentalkart", variant_spec=vspec,
    )
    return dk_match, dk_record


async def _resolve_by_code(
    dk_raw: list[CompetitorProduct], in_code: str, input_name: str
) -> tuple[CompetitorMatch, ProductRecord] | None:
    """Find the Dentalkart product (or grouped child) carrying the EXACT serial
    code the input names. Fetches the most plausible candidates' PDPs in
    parallel — codes live on the PDP / children, not the search card."""
    ranked = sorted(
        dk_raw,
        key=lambda c: fuzz_ratio(
            normalize_for_match(input_name), normalize_for_match(c.name)
        ),
        reverse=True,
    )[:8]
    pdps = await asyncio.gather(
        *(fetch_product("dentalkart", c.url) for c in ranked),
        return_exceptions=True,
    )
    for cand, pdp in zip(ranked, pdps, strict=True):
        rich = pdp if isinstance(pdp, CompetitorProduct) else cand
        children = rich.variants or []
        codes = [_paren_code(str(ch.get("name", ""))) for ch in children]
        # Only pick a child by code when children have DISTINCT codes. If they
        # all share one code (the parent SKU, e.g. "JULL-DENT 223"), the code
        # doesn't discriminate — defer to config/name resolution instead.
        if len({c for c in codes if c}) > 1:
            # Several children can share one code (DK data quirk — e.g. both
            # "Micro Forcep Tooth - Angled (041D)" and "Diamond Dusted … Angled
            # 45 (041D)"). Among code matches, pick the best NAME match to input.
            coded = [ch for ch, code in zip(children, codes, strict=True) if code == in_code]
            if coded:
                in_norm = normalize_for_match(input_name)
                best = max(
                    coded,
                    key=lambda ch: fuzz_ratio(in_norm, normalize_for_match(str(ch.get("name", "")))),
                )
                return _build_dk_result(rich, best, input_name)
        # A simple product whose own name carries the exact code.
        if not children and _paren_code(rich.name) == in_code:
            return _build_dk_result(rich, None, input_name)
    return None


async def _resolve_by_child_name(
    dk_raw: list[CompetitorProduct], input_name: str, current_fuzz: float
) -> tuple[CompetitorMatch, ProductRecord] | None:
    """The input may be a CHILD name while a DIFFERENT product out-ranked the
    correct grouped parent (e.g. "Julldent Implant Drivers and Hex Drivers"
    beats "Julldent Prosthetic Hex Drivers - Long", whose child IS the input).
    Scan the top candidates' children for a near-exact name match to the input
    and adopt it when it clearly beats the current (weak) resolution."""
    in_norm = normalize_for_match(input_name)
    ranked = sorted(
        dk_raw,
        key=lambda c: fuzz_ratio(in_norm, normalize_for_match(c.name)),
        reverse=True,
    )[:8]
    pdps = await asyncio.gather(
        *(fetch_product("dentalkart", c.url) for c in ranked),
        return_exceptions=True,
    )
    best_f = max(current_fuzz, 0.9)  # require near-exact AND better than current
    best: tuple[CompetitorProduct, dict] | None = None
    for cand, pdp in zip(ranked, pdps, strict=True):
        rich = pdp if isinstance(pdp, CompetitorProduct) else cand
        for ch in rich.variants or []:
            cn = normalize_for_match(str(ch.get("name", "")))
            # Respect hard gates (e.g. model-code mismatch): don't adopt a child
            # whose code conflicts with the input (Meril #2-0 must not take #3-0).
            if not gate_check(in_norm, cn).passed:
                continue
            f = fuzz_ratio(in_norm, cn)
            if f > best_f:
                best_f, best = f, (rich, ch)
    return _build_dk_result(best[0], best[1], input_name) if best else None


async def _pooled_dk_search(name: str) -> list[CompetitorProduct]:
    """DK's on-site search misses the grouped parent for a full child-name query
    but finds it for the config/size-stripped base name (e.g. the Orringer
    retractor parent only surfaces for 'Julldent Orringer Retractor', not the
    full child name). So fire the raw name AND the base name in parallel and pool
    by URL. (Broad progressive queries were tried too but added noise on the
    UNANCHORED self-match — 'GC Gold Label 9' pulling in 'GC Gold Label Hybrid'.
    Competitors keep the progressive queries because they're matched against the
    DK anchor with strict gates, which filters that noise out.)"""
    queries = [q for q in dict.fromkeys([name, base_name(name)]) if q and len(q) >= 3]
    results = await asyncio.gather(
        *(scrape_competitor("dentalkart", q) for q in queries),
        return_exceptions=True,
    )
    # Dedup by URL, keeping the first (most-specific-query) occurrence.
    pooled: dict[str, CompetitorProduct] = {}
    for r in results:
        if isinstance(r, list):
            for c in r:
                if c.url and c.url not in pooled:
                    pooled[c.url] = c
    return list(pooled.values())


async def _resolve_dk(row: DkRow) -> tuple[CompetitorMatch | None, ProductRecord | None]:
    """Search dentalkart.com, pick the best self-match, enrich via PDP."""
    dk_raw = await _pooled_dk_search(row.name)
    if not dk_raw:
        return None, None

    # Exact serial/model-code wins: when the input names a code (e.g.
    # "(5527/002/E)", "(KGF 9)"), the right product is the one carrying that
    # EXACT code — even if a different variant name-matches better (e.g. the
    # standalone "70 Microns - Red" outscores the grouped "Blue & Red" parent
    # whose name lacks "70 Microns"). Resolve by code first; fall through if no
    # exact match is found.
    in_code = _paren_code(row.name)
    if dk_raw and _looks_like_code(in_code):
        coded = await _resolve_by_code(dk_raw, in_code, row.name)
        if coded is not None:
            return coded

    dk_match = _best_match(row.name, dk_raw, None, qualifier_ref=row.name)
    if dk_match is None:
        # No parent name-matched — but the input may be a CHILD whose grouped
        # parent has a divergent name (e.g. input "…Suture Corn Pliers - Large"
        # under parent "Julldent Micro Tissue …Forcep (JULL-DENT 074)"). Scan the
        # top candidates' children for a near-exact match before giving up.
        better = await _resolve_by_child_name(dk_raw, row.name, 0.0)
        return better if better is not None else (None, None)
    dk_match.competitor_id = "dentalkart"
    dk_match.competitor_name = "Dentalkart"

    # When the top product is only a weak name match to the input, the input may
    # actually name a CHILD of a lower-ranked grouped product. Look for a
    # near-exact child across the top candidates before committing.
    cur_fuzz = fuzz_ratio(
        normalize_for_match(row.name), normalize_for_match(dk_match.matched_name or "")
    )
    if cur_fuzz < 0.9:
        better = await _resolve_by_child_name(dk_raw, row.name, cur_fuzz)
        if better is not None:
            return better

    pdp = await fetch_product("dentalkart", dk_match.matched_url or "")
    search_cand = next((c for c in dk_raw if c.url == dk_match.matched_url), None)
    src = pdp or search_cand
    if src is None:
        return dk_match, None
    dk_record = pipeline.record_from(src)

    # Grouped product → resolve to the exact child the input names, showing that
    # child's full name + its real price + stock (the parent listing collapses
    # every variant to one wrong price, e.g. Julldent → ₹1995 'Box Only').
    child = (
        _pick_dk_child(row.name, dk_match.matched_name or "", pdp.variants)
        if pdp and pdp.variants
        else None
    )
    if child is not None:
        dk_match.matched_name = str(child.get("name") or dk_match.matched_name)
        dk_match.matched_price = float(child.get("price") or dk_match.matched_price or 0)
        if child.get("image"):
            dk_match.matched_image = str(child["image"])
        if "inStock" in child:
            dk_match.in_stock = bool(child["inStock"])
        dk_record.name = dk_match.matched_name
        dk_record.price = dk_match.matched_price
        dk_record.pack_size = int(child.get("packSize") or 1)
        dk_record.unit_price = float(child.get("unitPrice") or dk_record.price)
        dk_record.variant_spec = VariantSpec.from_dict(child.get("variantSpec"))
        # The self-match score came from the grouped PARENT name (which lacks the
        # child's descriptors), so it reads low/"possible". Re-score against the
        # resolved child — an exact child match should read confirmed/high.
        rescored = triage_batch(row.name, [dk_match.matched_name])
        if rescored:
            dk_match.verdict = rescored[0].verdict.value
            dk_match.score = rescored[0].score
            dk_match.cosine = rescored[0].cosine
        return dk_match, dk_record

    # Dentalkart is the source of truth. Use the listing price and the grouped
    # parent's composition spec from the search API — the PDP often resolves to
    # a single cheaper child (e.g. GC Gold Label 9 PDP = ₹1369 / 5g, but the
    # listing is ₹2760 / 15g+13.1g).
    if dk_match.matched_price:
        dk_record.price = dk_match.matched_price
        if dk_record.pack_size and dk_record.pack_size > 1:
            dk_record.unit_price = round(dk_record.price / dk_record.pack_size, 2)
        else:
            dk_record.unit_price = dk_record.price
    if search_cand is not None and search_cand.variant_spec:
        dk_record.variant_spec = VariantSpec.from_dict(search_cand.variant_spec)
    return dk_match, dk_record


def _dk_has_input_product(input_name: str, dk_record: ProductRecord | None) -> bool:
    """True when Dentalkart actually carries the INPUT product (not just a
    different variant). If the input names a distinctive code/size (suture
    '#2-0', model 'DL-300', dims '17x25') that the DK match doesn't share, DK
    resolved to something else (or the product is delisted) — so competitors
    should be matched to the INPUT, not DK's wrong resolution."""
    if dk_record is None:
        return False
    in_attrs = extract_attributes(normalize_for_match(input_name))
    dk_attrs = extract_attributes(normalize_for_match(dk_record.name))
    in_codes = set(in_attrs.model_codes)
    if in_codes and not (in_codes & set(dk_attrs.model_codes)):
        return False
    # Upper vs Lower archwire is a different product. DK lists each separately
    # and its search returns the nearest sibling — don't let "Lower 016 X 022"
    # anchor on DK's "Upper 016 X 022" (then match competitors to the input).
    if in_attrs.wire_form and dk_attrs.wire_form and in_attrs.wire_form != dk_attrs.wire_form:
        return False
    return True


async def _compare_one(
    row: DkRow, db: Database | None, budget: JudgeBudget
) -> CompareResult:
    dk_match, dk_record = await _resolve_dk(row)

    # The INPUT is the source of truth for what the user wants. When DK carries
    # that exact product, anchor on its rich PDP record. When it doesn't
    # (delisted, or DK resolved to a different variant), match competitors
    # against the INPUT itself — otherwise a valid competitor result (e.g.
    # Pinkblue's "Meril Filasilk #2-0", which DK no longer stocks) is blocked by
    # DK's wrong anchor and the whole row comes back empty.
    if _dk_has_input_product(row.name, dk_record) and dk_match is not None:
        ref = dk_record
        dk_out: CompetitorMatch | None = dk_match
        dk_price = dk_match.matched_price
        dk_unit_price = dk_record.unit_price or dk_record.price
    else:
        ref = ProductRecord(name=row.name)
        dk_out = None  # DK doesn't carry the exact input product
        dk_price = None
        dk_unit_price = None

    product_id: int | None = None
    if db is not None and dk_out is not None and dk_record is not None:
        try:
            product_id = await registry.upsert_product(db, dk_record)
        except Exception:  # registry is best-effort
            product_id = None

    ctx = ProductContext(
        description=ref.description or None,
        packaging=ref.packaging or None,
        sku=ref.sku,
    )
    queries = extract_smart_queries(ref.name, ctx) or [row.name]

    async def one_competitor(cid: str, cname: str) -> CompetitorMatch:
        # Phase 2: registry hit -> cheap refresh.
        if db is not None and product_id is not None:
            try:
                links = await registry.get_active_links(db, product_id, cid)
            except Exception:  # registry is best-effort
                links = []
            # Refresh only settled links. A 'possible' (judge off/over
            # budget) must go through discovery again so it can be
            # re-judged instead of being frozen forever.
            if links and (
                links[0].status == "human_verified"
                or links[0].verdict in ("confirmed", "variant")
            ):
                cell = await pipeline.refresh(cid, links[0], ref)
                if cell is not None:
                    return _cell_to_match(cid, cname, cell, dk_price, dk_unit_price, ref)
        # Phase 1: full discovery (matched against `ref` — the DK product when DK
        # carries it, otherwise the user's input).
        cell = await pipeline.discover(
            cid, queries, ref,
            budget=budget, db=db, product_id=product_id, dk_price=dk_price,
        )
        return _cell_to_match(cid, cname, cell, dk_price, dk_unit_price, ref)

    out = list(await asyncio.gather(
        *(one_competitor(cid, cname) for cid, cname in COMPETITORS)
    ))
    return CompareResult(dentalkart=row, dentalkart_match=dk_out, competitors=out)


@router.post("/single", response_model=CompareResult)
async def compare_single(row: DkRow) -> CompareResult:
    db: Database | None
    try:
        db = await get_db()
    except Exception:  # run stateless without a DB
        db = None
    try:
        budget = JudgeBudget(get_settings().llm_judge_budget_per_run)
        return await _compare_one(row, db, budget)
    finally:
        if db is not None:
            await db.close()


_NAME_HEADERS = {"product name", "name", "product", "title"}


def _parse_dk_xlsx(content: bytes) -> list[DkRow]:
    """Pull just the product-name column out of the xlsx. Everything else
    (SKU, brand, price) is ignored — we derive what we need from the live
    dentalkart.com scrape."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    iter_rows = ws.iter_rows(values_only=True)
    try:
        header = next(iter_rows)
    except StopIteration:
        return []

    headers_lc = [str(h).strip().lower() if h is not None else "" for h in header]

    n: int | None = None
    for i, h in enumerate(headers_lc):
        if h in _NAME_HEADERS:
            n = i
            break
    if n is None:
        if len(headers_lc) == 1:
            n = 0
        else:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"no product-name column found. Looked for one of "
                    f"{sorted(_NAME_HEADERS)}; got headers {headers_lc}"
                ),
            )

    rows: list[DkRow] = []
    for r in iter_rows:
        if not r or n >= len(r) or r[n] is None:
            continue
        name = str(r[n]).strip()
        if name:
            rows.append(DkRow(name=name))
    return rows


def _name_column(headers_lc: list[str]) -> int | None:
    return next((i for i, h in enumerate(headers_lc) if h in _NAME_HEADERS), None)


def _parse_dk_csv(content: bytes) -> list[DkRow]:
    """Pull the product-name column out of a CSV. Same column rules as the xlsx
    parser; a single-column file with no recognized header is treated as a bare
    list of product names (first row included)."""
    text = content.decode("utf-8-sig", errors="replace")  # tolerate a BOM
    all_rows = [r for r in csv.reader(io.StringIO(text)) if any((c or "").strip() for c in r)]
    if not all_rows:
        return []

    headers_lc = [str(h).strip().lower() for h in all_rows[0]]
    n = _name_column(headers_lc)
    if n is not None:
        data = all_rows[1:]
    elif len(headers_lc) == 1:
        n, data = 0, all_rows  # bare single-column list — no header to skip
    else:
        raise HTTPException(
            status_code=400,
            detail=(
                f"no product-name column found. Looked for one of "
                f"{sorted(_NAME_HEADERS)}; got headers {headers_lc}"
            ),
        )

    rows: list[DkRow] = []
    for r in data:
        if n < len(r):
            name = str(r[n]).strip()
            if name:
                rows.append(DkRow(name=name))
    return rows


def _parse_dk_upload(filename: str, content: bytes) -> list[DkRow]:
    """Parse an uploaded product list as CSV or xlsx. Routes by file extension,
    falling back to the magic bytes (xlsx is a ZIP, starts with 'PK')."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _parse_dk_csv(content)
    if name.endswith((".xlsx", ".xlsm")) or content[:2] == b"PK":
        return _parse_dk_xlsx(content)
    # Unknown extension and not a zip → assume delimited text.
    return _parse_dk_csv(content)


@router.post("/batch", response_model=CompareBatchResponse)
async def compare_batch(
    file: UploadFile = File(...),
    concurrency: int = 2,
) -> CompareBatchResponse:
    """Run /compare/single for every row of the xlsx.

    Concurrency is bounded — each row already fans out across the configured
    competitors,
    so running too many rows in parallel will trip rate limits on competitor
    sites. Default 2 in-flight rows.
    """
    content = await file.read()
    rows = _parse_dk_upload(file.filename or "", content)
    if not rows:
        raise HTTPException(status_code=400, detail="no usable rows in the uploaded file")

    db: Database | None
    try:
        db = await get_db()
    except Exception:  # run stateless without a DB
        db = None
    budget = JudgeBudget(get_settings().llm_judge_budget_per_run)
    sem = asyncio.Semaphore(max(1, concurrency))

    async def gated(r: DkRow) -> CompareResult:
        async with sem:
            return await _compare_one(r, db, budget)

    try:
        results = await asyncio.gather(*(gated(r) for r in rows))
    finally:
        if db is not None:
            await db.close()
    return CompareBatchResponse(total=len(results), results=results)


@router.post("/batch-stream")
async def compare_batch_stream(
    file: UploadFile = File(...),
    concurrency: int = 2,
) -> StreamingResponse:
    """Same as /batch, but streams NDJSON progress so the UI can show
    'searched X of N' live. Emits one line per event:
      {"type":"start","total":N}
      {"type":"result","index":i,"done":k,"total":N,"result":{...}}  (completion order)
      {"type":"done","total":N}
    """
    content = await file.read()
    rows = _parse_dk_upload(file.filename or "", content)
    if not rows:
        raise HTTPException(status_code=400, detail="no usable rows in the uploaded file")

    async def stream() -> AsyncIterator[str]:
        yield json.dumps({"type": "start", "total": len(rows)}) + "\n"

        db: Database | None
        try:
            db = await get_db()
        except Exception:  # run stateless without a DB
            db = None
        budget = JudgeBudget(get_settings().llm_judge_budget_per_run)
        sem = asyncio.Semaphore(max(1, concurrency))

        async def gated(i: int, r: DkRow) -> tuple[int, CompareResult]:
            async with sem:
                return i, await _compare_one(r, db, budget)

        tasks = [asyncio.create_task(gated(i, r)) for i, r in enumerate(rows)]
        done = 0
        try:
            for fut in asyncio.as_completed(tasks):
                i, res = await fut
                done += 1
                yield json.dumps({
                    "type": "result", "index": i, "done": done,
                    "total": len(rows), "result": res.model_dump(),
                }) + "\n"
        finally:
            for t in tasks:
                t.cancel()
            if db is not None:
                await db.close()
        yield json.dumps({"type": "done", "total": len(rows)}) + "\n"

    return StreamingResponse(stream(), media_type="application/x-ndjson")
